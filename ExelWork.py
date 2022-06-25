from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook('razm(1).xlsx')
ws = wb['Title']
ws.title = 'Title'

class ExelWorker:
    Tags_Dict = {}

    def loadExelBook(self,bookName):
        workBook = load_workbook(bookName)
        return workBook

    def loadSheetInBook(self,sheetName,workBook):
        workSheet = workBook[sheetName]
        return workSheet

    #def getNamesByTags(self,workSheet,tag1=None,Tag2=None,Tag3=None):


    def getAllRows(self,workSheet):
        allValues = workSheet.iter_rows()
        return allValues



    def listRows(self,selectedAllRows):
        listedValues = list(selectedAllRows)
        return listedValues


    def getTagsRow(self,workSheet):
        tagsRow = workSheet['B1:1']
        return tagsRow


    def dictTagsValues(self, row):
        for cell in row:
            value = self.getCellValue(cell)
            ExelWorker.Tags_Dict.update({value: cell})


    def getCellValue(self,cell):
        value = cell.value
        return value


    def updateTags(self):
        workbook = self.loadExelBook('razm.xlsx')
        worksheet = self.loadSheetInBook('Sheet1', workbook)
        tagsRow = self.getTagsRow(worksheet)
        self.dictTagsValues(tagsRow)
        print('Updated')


Worker = ExelWorker()
workbook = Worker.loadExelBook('razm.xlsx')
worksheet = Worker.loadSheetInBook('Sheet1',workbook)
tagsRow = Worker.getTagsRow(worksheet)
Worker.dictTagsValues(tagsRow)
print(ExelWorker.Tags_Dict)
Worker.updateTags()



